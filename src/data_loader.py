import os
import pandas as pd
import json
from openpyxl.utils import get_column_letter
from datetime import datetime

def load_excel(file_path):
    """
    Loads the Excel file into a DataFrame.
    
    Expected Excel structure:
      - Row 0: The first cell is unused or a header; cells 1...n contain interview IDs.
      - Rows 1...m: Column 0 contains the question text; columns 1...n contain responses.
    """
    return pd.read_excel(file_path, header=None)


def save_outputs(aggregated_codes, interview_codes, model):
    """
    Generates both JSON and Excel output files for qualitative data coding.
    - Keeps the original order of questions.
    - Adds a "Count" column to indicate how many times each code appears across interviews.
    - Orders codes in descending frequency for each question.
    - Merges cells in the "Question" column for better readability.
    """

    # Ensure the output directory exists
    output_dir = "data/outputs"
    os.makedirs(output_dir, exist_ok=True)

    # Create a unique suffix based on current datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Define unique file names
    agg_codes_file = f"{output_dir}/{timestamp}_{model}_aggregated_codes.json"
    interview_codes_file = f"{output_dir}/{timestamp}_{model}_interview_codes.json"
    excel_file = f"{output_dir}/{timestamp}_{model}_coding_matrix.xlsx"

    # Save JSON Outputs
    with open(agg_codes_file, "w") as f:
        json.dump(aggregated_codes, f, indent=4)
    with open(interview_codes_file, "w") as f:
        json.dump(interview_codes, f, indent=4)

    # Extract interview IDs from interview_codes JSON
    interview_ids = list(interview_codes.keys())

    # Preserve original question order
    question_order = list(aggregated_codes.keys())

    # Prepare data: one row per (question, code) pair.
    data = []
    for question in question_order:  # Ensure original order
        all_codes = aggregated_codes[question]
        code_counts = []
        
        for code in all_codes:
            row = {"Question": question, "Code": code}
            count = 0  # Initialize count
            
            for interview in interview_ids:
                identified_codes = interview_codes.get(interview, {}).get(question, [])
                present = 1 if code in identified_codes else 0
                row[interview] = present
                count += present  # Increase count for each occurrence
            
            row["Count"] = count  # Store total occurrences
            code_counts.append(row)

        # Sort codes by frequency (descending) for the current question
        code_counts.sort(key=lambda x: x["Count"], reverse=True)
        data.extend(code_counts)

    # Convert data to a DataFrame
    df = pd.DataFrame(data)

    # Write DataFrame to Excel using openpyxl for formatting
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="CodingMatrix")
        workbook = writer.book
        worksheet = writer.sheets["CodingMatrix"]

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # Merge cells in the "Question" column for rows with the same question
        start_row = 2
        current_question = worksheet.cell(row=start_row, column=1).value

        for row in range(3, worksheet.max_row + 2):
            cell_value = worksheet.cell(row=row, column=1).value if row <= worksheet.max_row else None

            if cell_value != current_question:
                if row - 1 > start_row:
                    worksheet.merge_cells(
                        start_row=start_row,
                        start_column=1,
                        end_row=row - 1,
                        end_column=1
                    )
                current_question = cell_value
                start_row = row

    print(f"JSON files saved as: {agg_codes_file} and {interview_codes_file}")
    print(f"Excel file saved as: {excel_file}")
